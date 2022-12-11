# Librerías
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,PatternFill,Color
from openpyxl.styles.borders import Border,Side

#Función de estilo
def cellStyle(hoja,col,ren,val,mergeCell=False,color="FFFFFF",aline=False,bold=False):
    # Tamaño y fuente de letra
    ft = Font(name="Arial Narrow",size=10,bold=bold)
    # Alineación
    ali = Alignment(horizontal="center")
    # Bordes
    borde = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    # Color de fondo
    bgFill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color(color))
    
    # Asignar valores y estilo
    hoja.cell(column=col,row=ren,value=val).font = ft
    hoja.cell(column=col,row=ren).border = borde
    hoja.cell(column=col,row=ren).fill = bgFill

    # Combinarse o no combinarse
    if mergeCell != False:
        hoja.merge_cells(start_row=ren,start_column=col,end_row=ren,end_column=mergeCell)

    # Alinear o no alinear
    if aline:
        hoja.cell(column=col,row=ren).alignment = ali
